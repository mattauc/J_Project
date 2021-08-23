import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import json
import openpyxl

def modify_value(cell, data_value, key):

    profile = open('config.json')
    data = json.load(profile)
    
    value = data['profiles'][key].split(",")
    if len(value) > 2 and str(data_value).replace('.','',1).isdigit():
            
        fnt = '00000000'
        if float(data_value) > float(value[1]):
            #Make red
            fnt = Font(color="00FF0000")
        elif float(data_value) < float(value[0]):
            #Make blue
            fnt = Font(color="000000FF")
        cell.font = fnt

    profile.close()
    return

def update_column(report_info, data_values, worksheet):

    col = worksheet.max_column
    #Insert the date to the top row, furthest column
    date_cell = worksheet.cell(row = 1, column = col+1)
    date_cell.value = report_info[1]

    for count, key in enumerate(data_values):

        #Iterates down the rows, inserting value to match key
        cell = worksheet.cell(row = count+2, column = col+1)
        cell.value = data_values[key]

        modify_value(cell, data_values[key], key)

    return True

def create_sheet(report_info, data_values, workbook):

    try:
        #Attempt to open worksheet
        worksheet = workbook.get_sheet_by_name(report_info[0])
    except KeyError: 
        #Creates new worksheet if none present
        worksheet = workbook.create_sheet()
        worksheet.title = report_info[0]
        
        profile = open('config.json')
        data = json.load(profile)

        #Initializes row profiles
        for count, key in enumerate(data_values):
            cell = worksheet.cell(row = count+2, column = 1)

            value = data['profiles'][key].split(",")
            if len(value) > 2:
                cell.value = f"{key}   ({value[0]} - {value[1]})  {value[2]}"
            else:
                cell.value = f"{key}   {data['profiles'][key]}"

        profile.close()
    return worksheet

def create_workbook():

    try:
        #Attempts to open existing workbook
        workbook = load_workbook(r'Remdesvir_trial.xlsx')
    except FileNotFoundError:
        #Creates workbook if non exist
        workbook = xlsxwriter.Workbook(r'Remdesvir_trial.xlsx')
        workbook.close()

        #Loading existing workbook and removing initial worksheet
        workbook = load_workbook(r'Remdesvir_trial.xlsx')
        workbook.remove(workbook.worksheets[0])

    return workbook